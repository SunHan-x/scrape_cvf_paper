"""
Excel formatting utilities for CVF paper scraper.
Provides customizable column widths, text wrapping, and alignment.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Default column configurations
DEFAULT_COLUMN_CONFIG = {
    "Conference": {
        "width": 12,
        "wrap_text": False,
        "alignment": "center"
    },
    "Year": {
        "width": 8,
        "wrap_text": False,
        "alignment": "center"
    },
    "Title": {
        "width": 60,
        "wrap_text": True,
        "alignment": "left"
    },
    "Abstract": {
        "width": 80,
        "wrap_text": True,
        "alignment": "left"
    },
    "URL": {
        "width": 50,
        "wrap_text": False,
        "alignment": "left"
    },
    "PDF_URL": {
        "width": 50,
        "wrap_text": False,
        "alignment": "left"
    },
    "PDF_Path": {
        "width": 60,
        "wrap_text": False,
        "alignment": "left"
    }
}

def save_papers_to_excel(output_file, papers_by_sheet, column_config=None):
    """
    Save papers to Excel with custom formatting.
    
    Args:
        output_file: Path to output Excel file
        papers_by_sheet: Dictionary of {sheet_name: [paper_dicts]}
        column_config: Optional dictionary of column configurations
                      Format: {column_name: {width: int, wrap_text: bool, alignment: str}}
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Use default config if none provided
        if column_config is None:
            column_config = DEFAULT_COLUMN_CONFIG
        
        # Write data to Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, papers in papers_by_sheet.items():
                df = pd.DataFrame(papers)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Apply formatting
        format_excel_file(output_file, column_config)
        return True
        
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return False

def format_excel_file(excel_file, column_config=None):
    """
    Apply formatting to an existing Excel file.
    
    Args:
        excel_file: Path to Excel file
        column_config: Dictionary of column configurations
    """
    if column_config is None:
        column_config = DEFAULT_COLUMN_CONFIG
    
    # Load workbook
    wb = load_workbook(excel_file)
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get header row
        headers = [cell.value for cell in ws[1]]
        
        # Apply formatting to each column
        for col_idx, header in enumerate(headers, start=1):
            if header in column_config:
                config = column_config[header]
                col_letter = get_column_letter(col_idx)
                
                # Set column width
                if "width" in config:
                    ws.column_dimensions[col_letter].width = config["width"]
                
                # Get alignment setting
                horizontal = config.get("alignment", "left")
                wrap_text = config.get("wrap_text", False)
                
                # Apply to all cells in column (including header)
                for row in range(1, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row}"]
                    
                    # Apply alignment and wrap text
                    cell.alignment = Alignment(
                        horizontal=horizontal,
                        vertical="top",
                        wrap_text=wrap_text
                    )
                    
                    # Make header bold
                    if row == 1:
                        cell.font = Font(bold=True)
        
        # Set default row height for better readability with wrapped text
        ws.row_dimensions[1].height = 20  # Header row
        for row in range(2, ws.max_row + 1):
            # Auto-adjust row height for wrapped text
            # Default height is good for most cases
            pass
    
    # Save workbook
    wb.save(excel_file)

def create_custom_column_config(overrides=None):
    """
    Create a custom column configuration based on defaults with overrides.
    
    Args:
        overrides: Dictionary of column configurations to override defaults
                  Format: {column_name: {width: int, wrap_text: bool, alignment: str}}
    
    Returns:
        Complete column configuration dictionary
    
    Example:
        config = create_custom_column_config({
            "Title": {"width": 80, "wrap_text": True, "alignment": "left"},
            "Year": {"width": 10, "alignment": "center"}
        })
    """
    config = DEFAULT_COLUMN_CONFIG.copy()
    
    if overrides:
        for col_name, col_settings in overrides.items():
            if col_name in config:
                config[col_name].update(col_settings)
            else:
                config[col_name] = col_settings
    
    return config

# Alignment options for reference:
# - "left"
# - "center" 
# - "right"
# - "justify"
# - "distributed"
